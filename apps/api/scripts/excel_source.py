# apps/api/scripts/excel_source.py
# Shared reader for the LRC's per-college book-data Excel files. Each file
# has one sheet per program, all sharing the same column layout (with minor
# header spelling drift between files, e.g. "Image" vs "Images").

from pathlib import Path

import openpyxl

# canonical column name -> acceptable header spellings seen across files
COLUMN_ALIASES = {
    "Author(s)": ["Author(s)"],
    "Author(s) Full Name": ["Author(s) Full Name"],
    "Year": ["Year"],
    "Title": ["Title"],
    "Place of Publication": ["Place of Publication"],
    "Call No.": ["Call No."],
    "Book ID (Accession No.)": ["Book ID (Accession No.)"],
    "Description": ["Description"],
    "ISBN": ["ISBN"],
    "Images": ["Images", "Image"],
}
COLUMNS = list(COLUMN_ALIASES.keys())

def read_records(xlsx_path) -> list[dict]:
    """Returns one dict per real book row (has a Title), across every sheet
    in the file, with a 'program' key set to the sheet name."""
    wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in header:
                    col_idx[canonical] = header.index(alias)
                    break

        if "Title" not in col_idx:
            print(f"  WARNING: {sheet_name!r} in {xlsx_path} has no recognizable Title column, skipping sheet")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[col_idx["Title"]]
            if not title:
                continue
            record = {"program": sheet_name}
            for canonical in COLUMNS:
                value = row[col_idx[canonical]] if canonical in col_idx else None
                record[canonical] = value.strip() if isinstance(value, str) else value
            records.append(record)

    return records
